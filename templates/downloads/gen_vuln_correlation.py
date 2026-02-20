"""Generate vuln-correlation-template.xlsx with data validation dropdowns."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
section_font = Font(bold=True, size=13, color="1A365D")
subsection_font = Font(bold=True, size=11, color="1A365D")
note_font = Font(italic=True, size=10, color="718096")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align


def add_dv(ws, col, start, end, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(f"{col}{start}:{col}{end}")


# ============================================================
# Sheet 1: Asset Inventory
# ============================================================
ws = wb.active
ws.title = "Asset Inventory"

ws.merge_cells("A1:G1")
ws["A1"] = "Asset Inventory (from Module 2)"
ws["A1"].font = section_font
ws["A2"] = "List all products identified during attack surface discovery that require vulnerability correlation."
ws["A2"].font = note_font

headers = ["Asset/Hostname", "Product", "Version", "CPE 2.3 Identifier", "Internet-Exposed", "Service/Port", "Network Zone"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, 7)

# Example row
ws.cell(row=5, column=1, value="vpn.example.com")
ws.cell(row=5, column=2, value="FortiGate")
ws.cell(row=5, column=3, value="FortiOS 7.4.6")
ws.cell(row=5, column=4, value="cpe:2.3:o:fortinet:fortios:7.4.6:*:*:*:*:*:*:*")
style_data_rows(ws, 5, 25, 7)

add_dv(ws, "E", 5, 25, '"Yes,No"')
add_dv(ws, "G", 5, 25, '"Perimeter,DMZ,Internal,Cloud,Unknown"')

ws["A27"] = "CPE Validation: Verify AI-generated CPE strings at NVD CPE Search (nvd.nist.gov/products/cpe/search)."
ws["A27"].font = note_font

ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 48
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 14
ws.column_dimensions["G"].width = 16

# ============================================================
# Sheet 2: Vulnerability Correlation
# ============================================================
ws_vc = wb.create_sheet("Vulnerability Correlation")

ws_vc.merge_cells("A1:J1")
ws_vc["A1"] = "Vulnerability Correlation Table"
ws_vc["A1"].font = section_font

vc_headers = ["Asset", "CVE ID", "CVSS", "KEV Listed", "Remote Exploit", "Description", "Priority", "Vendor Advisory", "Status", "Notes"]
for i, h in enumerate(vc_headers, 1):
    ws_vc.cell(row=3, column=i, value=h)
style_header_row(ws_vc, 3, 10)
style_data_rows(ws_vc, 4, 35, 10)

add_dv(ws_vc, "D", 4, 35, '"Yes,No"')
add_dv(ws_vc, "E", 4, 35, '"Yes,No"')
add_dv(ws_vc, "G", 4, 35, '"P0,P1,P2,P3"')
add_dv(ws_vc, "I", 4, 35, '"Open,Mitigated,Patched,Accepted Risk"')

ws_vc.column_dimensions["A"].width = 22
ws_vc.column_dimensions["B"].width = 18
ws_vc.column_dimensions["C"].width = 8
ws_vc.column_dimensions["D"].width = 12
ws_vc.column_dimensions["E"].width = 14
ws_vc.column_dimensions["F"].width = 45
ws_vc.column_dimensions["G"].width = 10
ws_vc.column_dimensions["H"].width = 22
ws_vc.column_dimensions["I"].width = 16
ws_vc.column_dimensions["J"].width = 30

# ============================================================
# Sheet 3: Priority Framework
# ============================================================
ws_pf = wb.create_sheet("Priority Framework")

ws_pf.merge_cells("A1:C1")
ws_pf["A1"] = "P0-P3 Priority Framework"
ws_pf["A1"].font = section_font

pf_headers = ["Priority", "Criteria", "Response Timeframe"]
for i, h in enumerate(pf_headers, 1):
    ws_pf.cell(row=3, column=i, value=h)
style_header_row(ws_pf, 3, 3)

pf_data = [
    ["P0 -- Active Exploitation", "Internet-exposed asset + confirmed active exploitation (CISA KEV, CISA ICS advisory, or vendor PSIRT alert) + remote or unauthenticated exploitation possible", "Immediate (hours)"],
    ["P1 -- Critical", "Internet-exposed asset + CVSS 9.0+ + no confirmed exploitation yet, or confirmed exploitation but requires authentication", "Urgent (48 hours)"],
    ["P2 -- High", "Internet-exposed asset + CVSS 7.0-8.9, or internal-only asset with confirmed exploitation", "Standard patch cycle (1-2 weeks)"],
    ["P3 -- Monitor", "Known vulnerability but no internet exposure, or low CVSS with no exploitation evidence", "Next maintenance window (30-90 days)"],
]
for r, row_data in enumerate(pf_data, 4):
    for c, val in enumerate(row_data, 1):
        ws_pf.cell(row=r, column=c, value=val)
style_data_rows(ws_pf, 4, 7, 3)

ws_pf.column_dimensions["A"].width = 26
ws_pf.column_dimensions["B"].width = 70
ws_pf.column_dimensions["C"].width = 30

# ============================================================
# Sheet 4: Vendor PSIRT Tracking
# ============================================================
ws_vp = wb.create_sheet("Vendor PSIRTs")

ws_vp.merge_cells("A1:E1")
ws_vp["A1"] = "Vendor PSIRT Tracking"
ws_vp["A1"].font = section_font

vp_headers = ["Vendor", "PSIRT URL", "Products in Inventory", "Last Checked", "Subscription Active"]
for i, h in enumerate(vp_headers, 1):
    ws_vp.cell(row=3, column=i, value=h)
style_header_row(ws_vp, 3, 5)

# Pre-populate common OT vendor PSIRTs
vp_data = [
    ["Fortinet", "fortiguard.fortinet.com/psirt", "", "", ""],
    ["Siemens", "cert-portal.siemens.com/productcert/csaf/", "", "", ""],
    ["Schneider Electric", "download.schneider-electric.com/files?p_Doc_Ref=SEVD", "", "", ""],
    ["Rockwell Automation", "rockwellautomation.custhelp.com", "", "", ""],
    ["ABB", "search.abb.com/library/Download.aspx?", "", "", ""],
]
for r, row_data in enumerate(vp_data, 4):
    for c, val in enumerate(row_data, 1):
        ws_vp.cell(row=r, column=c, value=val)
style_data_rows(ws_vp, 4, 15, 5)

add_dv(ws_vp, "E", 4, 15, '"Yes,No"')

ws_vp.column_dimensions["A"].width = 22
ws_vp.column_dimensions["B"].width = 48
ws_vp.column_dimensions["C"].width = 28
ws_vp.column_dimensions["D"].width = 14
ws_vp.column_dimensions["E"].width = 20

# ============================================================
# Sheet 5: Sources Checked
# ============================================================
ws_sc = wb.create_sheet("Sources Checked")

ws_sc.merge_cells("A1:D1")
ws_sc["A1"] = "Sources Checked"
ws_sc["A1"].font = section_font
ws_sc["A2"] = "For each correlation cycle, document which sources were queried."
ws_sc["A2"].font = note_font

sc_headers = ["Source", "URL", "Checked", "Findings"]
for i, h in enumerate(sc_headers, 1):
    ws_sc.cell(row=4, column=i, value=h)
style_header_row(ws_sc, 4, 4)

sc_data = [
    ["NVD", "nvd.nist.gov", "", ""],
    ["CISA KEV", "cisa.gov/known-exploited-vulnerabilities-catalog", "", ""],
    ["CISA ICS Advisories", "cisa.gov/news-events/cybersecurity-advisories", "", ""],
    ["ICS Advisory Project", "icsadvisoryproject.com", "", ""],
    ["Vendor PSIRTs", "[see Vendor PSIRTs sheet]", "", ""],
]
for r, row_data in enumerate(sc_data, 5):
    for c, val in enumerate(row_data, 1):
        ws_sc.cell(row=r, column=c, value=val)
style_data_rows(ws_sc, 5, 15, 4)

add_dv(ws_sc, "C", 5, 15, '"Yes,No"')

ws_sc.column_dimensions["A"].width = 22
ws_sc.column_dimensions["B"].width = 48
ws_sc.column_dimensions["C"].width = 10
ws_sc.column_dimensions["D"].width = 40

# ============================================================
# Sheet 6: Change Log
# ============================================================
ws_cl = wb.create_sheet("Change Log")

ws_cl.merge_cells("A1:C1")
ws_cl["A1"] = "Vulnerability Correlation Change Log"
ws_cl["A1"].font = section_font

cl_headers = ["Date", "Change", "Changed By"]
for i, h in enumerate(cl_headers, 1):
    ws_cl.cell(row=3, column=i, value=h)
style_header_row(ws_cl, 3, 3)

ws_cl.cell(row=4, column=1, value="[today]")
ws_cl.cell(row=4, column=2, value="Initial correlation completed")
ws_cl.cell(row=4, column=3, value="[name]")
style_data_rows(ws_cl, 4, 25, 3)

ws_cl.column_dimensions["A"].width = 14
ws_cl.column_dimensions["B"].width = 50
ws_cl.column_dimensions["C"].width = 18

wb.save("/home/cutaway/Projects/ot-osint-program-workshop/templates/downloads/vuln-correlation-template.xlsx")
print("vuln-correlation-template.xlsx created")
