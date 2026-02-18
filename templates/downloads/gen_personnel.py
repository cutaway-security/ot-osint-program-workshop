"""Generate personnel-inventory-template.xlsx with data validation dropdowns."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
section_font = Font(bold=True, size=13, color="1A365D")
subsection_font = Font(bold=True, size=11, color="1A365D")
note_font = Font(italic=True, size=10, color="718096")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align


def add_dv(ws, col, start, end, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    ws.add_data_validation(dv)
    dv.add(f"{col}{start}:{col}{end}")


# ============================================================
# Sheet 1: Email Pattern
# ============================================================
ws = wb.active
ws.title = "Email Pattern"

ws.merge_cells("A1:B1")
ws["A1"] = "Personnel Exposure Inventory -- Email Pattern"
ws["A1"].font = section_font

labels = [
    ("Organization:", "[name]"),
    ("Primary domain:", "[domain]"),
    ("Email format:", "[e.g., firstname.lastname@domain]"),
    ("Discovered at:", "[URL where pattern was confirmed]"),
]
for i, (label, val) in enumerate(labels, 3):
    ws.cell(row=i, column=1, value=label).font = Font(bold=True)
    ws.cell(row=i, column=2, value=val)

ws["A8"] = "Team Email Addresses"
ws["A8"].font = subsection_font

for i, h in enumerate(["Team Inbox", "Purpose"], 1):
    ws.cell(row=9, column=i, value=h)
style_header_row(ws, 9, 2)
style_data_rows(ws, 10, 18, 2)

ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 50

# ============================================================
# Sheet 2: Tier 1
# ============================================================
ws1 = wb.create_sheet("Tier 1 - Immediate")

ws1.merge_cells("A1:H1")
ws1["A1"] = "Tier 1 -- Immediate Priority"
ws1["A1"].font = section_font
ws1["A2"] = "Personnel with direct access to operational technology and critical control systems."
ws1["A2"].font = note_font

headers = ["Name", "Role", "Email", "Breach Status", "Breaches Found", "Data Exposed", "Priority", "Notes"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=4, column=i, value=h)
style_header_row(ws1, 4, 8)
style_data_rows(ws1, 5, 20, 8)

add_dv(ws1, "D", 5, 20, '"Clear,Breached"')
add_dv(ws1, "F", 5, 20, '"Email Only,Password Hash,Plaintext Password,Multiple Types"')
add_dv(ws1, "G", 5, 20, '"Critical,High,Monitor"')

for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws1.column_dimensions[c].width = 18
ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 32
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["E"].width = 16
ws1.column_dimensions["H"].width = 30

# ============================================================
# Sheet 3: Tier 2
# ============================================================
ws2 = wb.create_sheet("Tier 2 - High")

ws2.merge_cells("A1:H1")
ws2["A1"] = "Tier 2 -- High Priority"
ws2["A1"].font = section_font
ws2["A2"] = "Personnel with IT/OT bridge access or indirect OT system exposure."
ws2["A2"].font = note_font

for i, h in enumerate(headers, 1):
    ws2.cell(row=4, column=i, value=h)
style_header_row(ws2, 4, 8)
style_data_rows(ws2, 5, 20, 8)

add_dv(ws2, "D", 5, 20, '"Clear,Breached"')
add_dv(ws2, "F", 5, 20, '"Email Only,Password Hash,Plaintext Password,Multiple Types"')
add_dv(ws2, "G", 5, 20, '"Critical,High,Monitor"')

for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws2.column_dimensions[c].width = 18
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 30
ws2.column_dimensions["E"].width = 16
ws2.column_dimensions["H"].width = 30

# ============================================================
# Sheet 4: Tier 3
# ============================================================
ws3 = wb.create_sheet("Tier 3 - Elevated")

ws3.merge_cells("A1:H1")
ws3["A1"] = "Tier 3 -- Elevated Priority"
ws3["A1"].font = section_font
ws3["A2"] = "Executives and staff without direct system access but with organizational authority."
ws3["A2"].font = note_font

for i, h in enumerate(headers, 1):
    ws3.cell(row=4, column=i, value=h)
style_header_row(ws3, 4, 8)
style_data_rows(ws3, 5, 20, 8)

add_dv(ws3, "D", 5, 20, '"Clear,Breached"')
add_dv(ws3, "F", 5, 20, '"Email Only,Password Hash,Plaintext Password,Multiple Types"')
add_dv(ws3, "G", 5, 20, '"Critical,High,Monitor"')

for c in ["A", "B", "C", "D", "E", "F", "G", "H"]:
    ws3.column_dimensions[c].width = 18
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 32
ws3.column_dimensions["C"].width = 30
ws3.column_dimensions["E"].width = 16
ws3.column_dimensions["H"].width = 30

# ============================================================
# Sheet 5: Prioritization Matrix
# ============================================================
ws_pm = wb.create_sheet("Prioritization Matrix")

ws_pm.merge_cells("A1:D1")
ws_pm["A1"] = "Prioritization Matrix"
ws_pm["A1"].font = section_font

pm_headers = ["Tier", "Breach Severity", "Priority", "Action"]
for i, h in enumerate(pm_headers, 1):
    ws_pm.cell(row=3, column=i, value=h)
style_header_row(ws_pm, 3, 4)

pm_data = [
    ["Tier 1", "Password/hash exposed", "Critical", "Immediate credential reset. Verify MFA on all remote access. Review access logs."],
    ["Tier 1", "Email-only exposure", "High", "Monitor for credential stuffing. Verify MFA status. Add to enhanced monitoring."],
    ["Tier 1", "No breach found", "Monitor", "Baseline documented. Include in ongoing HIBP monitoring."],
    ["Tier 2", "Password/hash exposed", "High", "Credential reset. Review bridge access scope. Verify MFA."],
    ["Tier 2", "Email-only exposure", "Monitor", "Standard monitoring. Verify MFA on bridge accounts."],
    ["Tier 3", "Password/hash exposed", "High", "Credential reset. Social engineering awareness."],
    ["Tier 3", "Email-only exposure", "Monitor", "Standard monitoring."],
]
for r, row_data in enumerate(pm_data, 4):
    for c, val in enumerate(row_data, 1):
        ws_pm.cell(row=r, column=c, value=val)
style_data_rows(ws_pm, 4, 10, 4)

ws_pm.column_dimensions["A"].width = 12
ws_pm.column_dimensions["B"].width = 24
ws_pm.column_dimensions["C"].width = 14
ws_pm.column_dimensions["D"].width = 60

# ============================================================
# Sheet 6: Professional Network Exposure
# ============================================================
ws_pn = wb.create_sheet("Professional Network")

ws_pn.merge_cells("A1:D1")
ws_pn["A1"] = "Professional Network Exposure"
ws_pn["A1"].font = section_font

pn_headers = ["Source", "Finding", "Risk", "Notes"]
for i, h in enumerate(pn_headers, 1):
    ws_pn.cell(row=3, column=i, value=h)
style_header_row(ws_pn, 3, 4)

pn_examples = [
    ["Job postings", "[technology details disclosed]", "[social engineering / recon value]", ""],
    ["Conference talks", "[operational details disclosed]", "", ""],
    ["Professional profiles", "[certifications / tools revealed]", "", ""],
]
for r, row_data in enumerate(pn_examples, 4):
    for c, val in enumerate(row_data, 1):
        ws_pn.cell(row=r, column=c, value=val)
style_data_rows(ws_pn, 4, 15, 4)

add_dv(ws_pn, "A", 4, 15, '"Job Postings,Conference Talks,Professional Profiles,Social Media,Published Papers,Other"')

ws_pn.column_dimensions["A"].width = 22
ws_pn.column_dimensions["B"].width = 40
ws_pn.column_dimensions["C"].width = 35
ws_pn.column_dimensions["D"].width = 30

# ============================================================
# Sheet 7: Change Log
# ============================================================
ws_cl = wb.create_sheet("Change Log")

ws_cl.merge_cells("A1:C1")
ws_cl["A1"] = "Personnel Exposure Inventory Change Log"
ws_cl["A1"].font = section_font

cl_headers = ["Date", "Change", "Changed By"]
for i, h in enumerate(cl_headers, 1):
    ws_cl.cell(row=3, column=i, value=h)
style_header_row(ws_cl, 3, 3)

ws_cl.cell(row=4, column=1, value="[today]")
ws_cl.cell(row=4, column=2, value="Initial inventory created")
ws_cl.cell(row=4, column=3, value="[name]")
style_data_rows(ws_cl, 4, 25, 3)

ws_cl.column_dimensions["A"].width = 14
ws_cl.column_dimensions["B"].width = 50
ws_cl.column_dimensions["C"].width = 18

wb.save("/home/cutaway/Projects/ot-osint-program-workshop/templates/downloads/personnel-inventory-template.xlsx")
print("personnel-inventory-template.xlsx created")
