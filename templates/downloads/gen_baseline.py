"""Generate baseline-template.xlsx with data validation dropdowns."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
section_font = Font(bold=True, size=13, color="1A365D")
subsection_font = Font(bold=True, size=11, color="1A365D")
note_font = Font(italic=True, size=10, color="718096")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = wrap_align


def add_data_validation(ws, col_letter, start_row, end_row, formula):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select from dropdown"
    dv.promptTitle = "Choose Value"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


# ============================================================
# Sheet 1: Classification Key
# ============================================================
ws_key = wb.active
ws_key.title = "Classification Key"

ws_key.merge_cells("A1:D1")
ws_key["A1"] = "Baseline Document -- Classification Key"
ws_key["A1"].font = section_font

headers = ["Classification", "Meaning", "Action"]
for i, h in enumerate(headers, 1):
    ws_key.cell(row=3, column=i, value=h)
style_header_row(ws_key, 3, 3)

data = [
    ["Known-Good", "Expected and properly configured. No security concern.", "Document and monitor for changes."],
    ["Accepted Risk", "Known exposure that cannot be remediated immediately.", "Document business justification. Monitor frequently. Review quarterly."],
    ["Needs Remediation", "Security issue with a clear fix available.", "Add to remediation queue with priority. Track to resolution."],
    ["Needs Investigation", "Unknown or unexpected item requiring analysis.", "Investigate within one monitoring cycle. Do not leave items here indefinitely."],
]
for r, row_data in enumerate(data, 4):
    for c, val in enumerate(row_data, 1):
        ws_key.cell(row=r, column=c, value=val)
style_data_rows(ws_key, 4, 7, 3)

ws_key.column_dimensions["A"].width = 22
ws_key.column_dimensions["B"].width = 55
ws_key.column_dimensions["C"].width = 55

# ============================================================
# Sheet 2: Attack Surface (M2)
# ============================================================
ws_as = wb.create_sheet("Attack Surface (M2)")

ws_as.merge_cells("A1:G1")
ws_as["A1"] = "Section 1: External Attack Surface (Module 2)"
ws_as["A1"].font = section_font

ws_as["A2"] = "List remote access assets first. These are the highest-priority entries."
ws_as["A2"].font = note_font

# Remote Access subsection
ws_as["A4"] = "Remote Access and Administrative Interfaces"
ws_as["A4"].font = subsection_font

ra_headers = ["Hostname/URL", "IP Address", "Port", "Product/Version", "Classification", "Priority", "Notes"]
for i, h in enumerate(ra_headers, 1):
    ws_as.cell(row=5, column=i, value=h)
style_header_row(ws_as, 5, 7)

# Example row
ws_as.cell(row=6, column=1, value="vpn.example.com")
ws_as.cell(row=6, column=2, value="x.x.x.x")
ws_as.cell(row=6, column=3, value="443")
ws_as.cell(row=6, column=4, value="FortiOS 7.4.6")
style_data_rows(ws_as, 6, 20, 7)

# Classification dropdown for remote access
add_data_validation(ws_as, "E", 6, 20, '"Known-Good,Accepted Risk,Needs Remediation,Needs Investigation"')
# Priority dropdown
add_data_validation(ws_as, "F", 6, 20, '"P0,P1,P2,P3"')

# Domains subsection
ws_as["A22"] = "Domains and Subdomains"
ws_as["A22"].font = subsection_font

dom_headers = ["Subdomain", "IP Address", "Category", "Classification", "Notes"]
for i, h in enumerate(dom_headers, 1):
    ws_as.cell(row=23, column=i, value=h)
style_header_row(ws_as, 23, 5)

# Example rows
ws_as.cell(row=24, column=1, value="mail.example.com")
ws_as.cell(row=24, column=2, value="x.x.x.x")
ws_as.cell(row=24, column=3, value="Email")
ws_as.cell(row=25, column=1, value="sso.example.com")
ws_as.cell(row=25, column=2, value="x.x.x.x")
ws_as.cell(row=25, column=3, value="Authentication")
style_data_rows(ws_as, 24, 50, 5)

# Category dropdown
add_data_validation(ws_as, "C", 24, 50, '"Email,Authentication,Web Application,Benefits/HR,Financial,Staging/Test,Infrastructure,Other"')
# Classification dropdown
add_data_validation(ws_as, "D", 24, 50, '"Known-Good,Accepted Risk,Needs Remediation,Needs Investigation"')

# Exposed Services subsection
ws_as["A52"] = "Exposed Services"
ws_as["A52"].font = subsection_font

svc_headers = ["IP Address", "Port", "Service", "Product/Version", "Classification", "Notes"]
for i, h in enumerate(svc_headers, 1):
    ws_as.cell(row=53, column=i, value=h)
style_header_row(ws_as, 53, 6)
style_data_rows(ws_as, 54, 70, 6)

add_data_validation(ws_as, "E", 54, 70, '"Known-Good,Accepted Risk,Needs Remediation,Needs Investigation"')

for col in ["A", "B", "C", "D", "E", "F", "G"]:
    ws_as.column_dimensions[col].width = 20
ws_as.column_dimensions["A"].width = 28
ws_as.column_dimensions["D"].width = 22
ws_as.column_dimensions["G"].width = 30

# ============================================================
# Sheet 3: Personnel Exposure (M3)
# ============================================================
ws_pe = wb.create_sheet("Personnel (M3)")

ws_pe.merge_cells("A1:G1")
ws_pe["A1"] = "Section 2: Personnel Exposure (Module 3)"
ws_pe["A1"].font = section_font

ws_pe["A3"] = "Email Format"
ws_pe["A3"].font = subsection_font
ws_pe["A4"] = "Primary domain:"
ws_pe["B4"] = "[domain]"
ws_pe["A5"] = "Email pattern:"
ws_pe["B5"] = "[firstname.lastname@domain]"
ws_pe["A6"] = "Team addresses:"
ws_pe["B6"] = "[list any shared/team inboxes]"
ws_pe["A4"].font = Font(bold=True)
ws_pe["A5"].font = Font(bold=True)
ws_pe["A6"].font = Font(bold=True)

ws_pe["A8"] = "Personnel Inventory"
ws_pe["A8"].font = subsection_font

pe_headers = ["Name", "Role", "Tier", "Email", "Breach Status", "Data Exposed", "Classification", "Priority", "Notes"]
for i, h in enumerate(pe_headers, 1):
    ws_pe.cell(row=9, column=i, value=h)
style_header_row(ws_pe, 9, 9)
style_data_rows(ws_pe, 10, 30, 9)

add_data_validation(ws_pe, "C", 10, 30, '"Tier 1,Tier 2,Tier 3"')
add_data_validation(ws_pe, "E", 10, 30, '"Clear,Breached"')
add_data_validation(ws_pe, "F", 10, 30, '"Email Only,Password Hash,Plaintext Password,Multiple Types"')
add_data_validation(ws_pe, "G", 10, 30, '"Known-Good,Accepted Risk,Needs Remediation,Needs Investigation"')
add_data_validation(ws_pe, "H", 10, 30, '"Critical,High,Monitor"')

for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
    ws_pe.column_dimensions[col].width = 18
ws_pe.column_dimensions["A"].width = 22
ws_pe.column_dimensions["B"].width = 30
ws_pe.column_dimensions["D"].width = 28
ws_pe.column_dimensions["I"].width = 30

# ============================================================
# Sheet 4: Vulnerability Correlation (M4)
# ============================================================
ws_vc = wb.create_sheet("Vulnerabilities (M4)")

ws_vc.merge_cells("A1:I1")
ws_vc["A1"] = "Section 3: Vulnerability Correlation (Module 4)"
ws_vc["A1"].font = section_font

ws_vc["A3"] = "Asset-Vulnerability Table"
ws_vc["A3"].font = subsection_font

vc_headers = ["Asset", "Product/Version", "CPE", "CVE(s)", "CVSS", "KEV Listed", "Priority", "Vendor Advisory", "Status", "Notes"]
for i, h in enumerate(vc_headers, 1):
    ws_vc.cell(row=4, column=i, value=h)
style_header_row(ws_vc, 4, 10)
style_data_rows(ws_vc, 5, 25, 10)

add_data_validation(ws_vc, "F", 5, 25, '"Yes,No"')
add_data_validation(ws_vc, "G", 5, 25, '"P0,P1,P2,P3"')
add_data_validation(ws_vc, "I", 5, 25, '"Open,Mitigated,Patched,Accepted Risk"')

for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
    ws_vc.column_dimensions[col].width = 18
ws_vc.column_dimensions["A"].width = 24
ws_vc.column_dimensions["B"].width = 22
ws_vc.column_dimensions["C"].width = 40
ws_vc.column_dimensions["D"].width = 18
ws_vc.column_dimensions["H"].width = 24
ws_vc.column_dimensions["J"].width = 30

# ============================================================
# Sheet 5: Monitoring Configuration (M5)
# ============================================================
ws_mc = wb.create_sheet("Monitoring Config (M5)")

ws_mc.merge_cells("A1:C1")
ws_mc["A1"] = "Section 4: Monitoring Configuration (Module 5)"
ws_mc["A1"].font = section_font

ws_mc["A3"] = "Push-Based Alert Sources"
ws_mc["A3"].font = subsection_font

push_headers = ["Service", "Configuration", "Status"]
for i, h in enumerate(push_headers, 1):
    ws_mc.cell(row=4, column=i, value=h)
style_header_row(ws_mc, 4, 3)

push_data = [
    ["Google Alerts", "[list queries configured]", ""],
    ["CISA ICS Advisories", "[subscription email]", ""],
    ["CISA KEV Updates", "[subscription email]", ""],
    ["Vendor PSIRTs", "[list vendors subscribed]", ""],
    ["HIBP Domain", "[domain if verified]", ""],
    ["Certificate Transparency", "[monitoring service]", ""],
    ["Shodan Monitor", "[IPs configured]", ""],
]
for r, row_data in enumerate(push_data, 5):
    for c, val in enumerate(row_data, 1):
        ws_mc.cell(row=r, column=c, value=val)
style_data_rows(ws_mc, 5, 15, 3)

add_data_validation(ws_mc, "C", 5, 15, '"Active,Inactive,Not Configured"')

ws_mc["A17"] = "Pull-Based Schedule Reference"
ws_mc["A17"].font = subsection_font
ws_mc["A18"] = "Weekly checks:"
ws_mc["B18"] = "[list from Artifact 6]"
ws_mc["A19"] = "Monthly checks:"
ws_mc["B19"] = "[list from Artifact 6]"
ws_mc["A20"] = "Schedule owner:"
ws_mc["B20"] = "[name]"
ws_mc["A21"] = "Backup owner:"
ws_mc["B21"] = "[name]"
ws_mc["A18"].font = Font(bold=True)
ws_mc["A19"].font = Font(bold=True)
ws_mc["A20"].font = Font(bold=True)
ws_mc["A21"].font = Font(bold=True)

ws_mc.column_dimensions["A"].width = 28
ws_mc.column_dimensions["B"].width = 40
ws_mc.column_dimensions["C"].width = 18

# ============================================================
# Sheet 6: Change Log
# ============================================================
ws_cl = wb.create_sheet("Change Log")

ws_cl.merge_cells("A1:D1")
ws_cl["A1"] = "Baseline Document Change Log"
ws_cl["A1"].font = section_font

cl_headers = ["Date", "Section", "Change Description", "Changed By"]
for i, h in enumerate(cl_headers, 1):
    ws_cl.cell(row=3, column=i, value=h)
style_header_row(ws_cl, 3, 4)

ws_cl.cell(row=4, column=1, value="[today]")
ws_cl.cell(row=4, column=2, value="All")
ws_cl.cell(row=4, column=3, value="Initial baseline created")
ws_cl.cell(row=4, column=4, value="[name]")
style_data_rows(ws_cl, 4, 30, 4)

add_data_validation(ws_cl, "B", 4, 30, '"All,Attack Surface (M2),Personnel (M3),Vulnerabilities (M4),Monitoring Config (M5)"')

ws_cl.column_dimensions["A"].width = 14
ws_cl.column_dimensions["B"].width = 24
ws_cl.column_dimensions["C"].width = 50
ws_cl.column_dimensions["D"].width = 18

# Save
wb.save("/home/cutaway/Projects/ot-osint-program-workshop/templates/downloads/baseline-template.xlsx")
print("baseline-template.xlsx created")
